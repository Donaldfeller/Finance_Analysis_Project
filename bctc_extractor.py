"""
BCTC Extractor - Trích xuất Báo cáo tài chính (PDF) bằng OpenRouter (openai/gpt-5.4-mini)
----------------------------------------------------------------------------------------
- Model đọc trực tiếp file PDF (gửi dưới dạng base64 "file" content-part) - KHÔNG qua OCR/
  bước trung gian nào khác.
- Người dùng có thể xếp hàng (queue) TỐI ĐA 10 file PDF. Các file được xử lý LẦN LƯỢT:
  file trước phải được trích xuất + lưu xong vào thư mục chỉ định thì file sau mới bắt đầu.
- Kết quả xuất ra 1 trong các định dạng phổ biến với data analyst (chọn bằng dropdown):
    - Excel (.xlsx)          -> 1 file, 3 sheet (BCĐKT / KQKD / LCTT)
    - SQLite Database (.db)  -> 1 file, 3 table
    - JSON (.json)           -> 1 file, 3 khối dữ liệu
    - CSV (.csv)             -> 3 file .csv riêng (do CSV không hỗ trợ nhiều bảng/sheet)
  Tên file (hoặc tiền tố tên file, đối với CSV) LUÔN trùng với tên file PDF gốc.
- Mỗi bảng CHỈ giữ phần CỐT LÕI: Mã số | Chỉ tiêu | Thuyết minh | các cột số liệu
  (không có tiêu đề báo cáo, không có tên doanh nghiệp/đơn vị tiền tệ/kỳ báo cáo...).
- Giao diện Tkinter tối giản: thêm/xoá PDF vào hàng đợi, chọn thư mục lưu, chọn định dạng
  xuất, nút "Bắt đầu xử lý", log tiến trình / lỗi hiển thị trực tiếp (không bị "skip" âm thầm).

Yêu cầu thư viện: requests, openpyxl (đã có sẵn trong máy). sqlite3/csv/json thuộc thư viện chuẩn.
"""

import base64
import csv
import json
import os
import re
import sqlite3
import tempfile
import threading
import traceback
from datetime import datetime

import requests
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ============================== THƯ VIỆN CHO GIẢI PHÁP SỬA TRANG BỊ XOAY NGANG ==============================
# BỐI CẢNH LỖI: PDF BCTC này là ẢNH SCAN THUẦN (đã kiểm chứng: page.get_text() trả về rỗng ở
# MỌI trang -> không có text layer để dựa vào). Một số trang (thường là Báo cáo KQKD hợp nhất
# theo QUÝ, do có 4 cột số liệu nên khi in/scan kế toán hay xoay ngang trang đó) bị nhúng vào
# PDF mà NỘI DUNG PIXEL bên trong đã xoay 90/180/270 độ so với chiều đọc đúng - trong khi
# MediaBox và kích thước ảnh nhúng vẫn là portrait giống mọi trang khác. Vì vậy:
#   - KHÔNG thể phát hiện bằng cách so sánh width/height (đã kiểm chứng: mọi trang, kể cả
#     trang lỗi, đều có MediaBox 595x842 và ảnh nhúng 1654x2340 - portrait y hệt nhau).
#   - KHÔNG thể tin vào pytesseract.image_to_osd(): đã thử và bị nhận NHẦM ngôn ngữ/kịch bản
#     (Cyrillic/Arabic...) do BCTC có rất nhiều số + ký hiệu bảng, độ tin cậy rất thấp.
#   - Dùng riêng OCR-confidence-score (thử 4 góc, chọn góc có tổng confidence cao nhất) CŨNG
#     KHÔNG đủ tin cậy: bị nhiễu bởi watermark hình logo, dấu mộc đỏ, chữ ký... khiến có lúc
#     chọn sai góc đối xứng (180° thay vì 90° đúng) trên chính trang lỗi.
#
# => GIẢI PHÁP 2 TẦNG (đã thực nghiệm và xác thực đúng trên toàn bộ các trang lỗi + trang OK):
#   TẦNG 1 - Projection Profile Variance (không dùng OCR, không bị ảnh hưởng bởi watermark):
#       Nhị phân hoá ảnh (pixel tối = chữ), tính tổng pixel chữ theo từng DÒNG NGANG (row),
#       rồi tính VARIANCE của các tổng dòng đó. Khi ảnh đang ở đúng hướng ĐỌC THEO CHIỀU NGANG
#       (chữ chạy dọc theo các dòng ngang), các dòng có chữ và các dòng trống xen kẽ rõ rệt
#       => variance cao. Khi ảnh bị xoay 90°/270° (chữ chạy dọc theo cột), việc chiếu ngang sẽ
#       "trộn" nhiều cột chữ+trống vào mỗi dòng => phân bố đều hơn => variance thấp.
#       Nhờ vậy, TẦNG 1 phân loại chính xác ảnh đang ở NHÓM GÓC nào: {0°, 180°} hay {90°, 270°},
#       hoàn toàn không cần OCR nên không bị nhiễu bởi watermark/dấu mộc/chữ ký.
#   TẦNG 2 - OCR confidence (Tesseract, lang=vie+eng) CHỈ dùng để phân biệt 2 lựa chọn CÙNG
#       NHÓM (0 vs 180, hoặc 90 vs 270). Vì chỉ so sánh 2 ứng viên đối lập nhau 180° (một
#       ĐÚNG hẳn, một NGƯỢC hẳn), độ nhiễu OCR không đủ để đánh lừa kết quả (thực nghiệm cho
#       thấy góc đúng luôn có điểm OCR cao hơn rõ rệt góc ngược 180° của nó).
try:
    import fitz  # PyMuPDF
    import numpy as np
    from PIL import Image
    _ROTATION_FIX_CORE_LIBS_AVAILABLE = True
except ImportError:
    _ROTATION_FIX_CORE_LIBS_AVAILABLE = False

try:
    import pytesseract
    _PYTESSERACT_AVAILABLE = True
except ImportError:
    _PYTESSERACT_AVAILABLE = False

# Ngôn ngữ OCR: tiếng Việt + tiếng Anh (BCTC có cả chữ Việt có dấu và số/ký hiệu Latin).
OCR_LANG = "vie+eng"

# DPI khi render trang PDF ra ảnh để phát hiện góc xoay (không cần cao vì chỉ dùng để chấm
# điểm hướng đọc, không ảnh hưởng chất lượng dữ liệu gửi cho model trích xuất).
OCR_DETECT_DPI = 120

# Ngưỡng an toàn cho TẦNG 1 (projection-variance): chỉ coi 1 trang là "cần xoay 90°/270°"
# nếu tỉ lệ variance(90°) / variance(0°) VƯỢT ngưỡng này. Lý do cần ngưỡng: các trang dạng
# VĂN BẢN THƯỜNG (không phải bảng số liệu, ví dụ trang bìa, trang thuyết minh) có variance
# giữa 2 hướng không chênh lệch nhiều do nội dung thưa/không đều => dễ bị nhiễu nếu so sánh
# tuyệt đối (0 >= 90 hay không). Thực nghiệm trên toàn bộ 33 trang của file mẫu cho thấy:
#   - Các trang THỰC SỰ bị xoay sai (dạng bảng số liệu) có ratio 90/0 nằm trong khoảng 3.69–7.34.
#   - TẤT CẢ các trang khác (kể cả trang dễ gây nhiễu nhất) có ratio chỉ tối đa 2.29.
# => Ngưỡng 3.0 nằm giữa khoảng trống an toàn (2.29 - 3.69), phân loại đúng 100% trên dữ liệu
#    thực nghiệm, đồng thời vẫn có biên độ an toàn cho các file PDF khác.
ROTATION_PROJECTION_RATIO_THRESHOLD = 3.0


# Các đường dẫn cài đặt Tesseract-OCR phổ biến trên Windows, dùng để tự động dò tìm khi
# tesseract.exe chưa nằm trong PATH của hệ thống.
_TESSERACT_CANDIDATE_PATHS = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    os.path.expandvars(r"%LOCALAPPDATA%\Programs\Tesseract-OCR\tesseract.exe"),
]


def _configure_tesseract() -> bool:
    """Đảm bảo pytesseract tìm được tesseract.exe. Trả về True nếu sẵn sàng dùng được OCR."""
    if not _PYTESSERACT_AVAILABLE:
        return False
    try:
        pytesseract.pytesseract.get_tesseract_version()
        return True
    except Exception:
        pass
    for candidate in _TESSERACT_CANDIDATE_PATHS:
        if candidate and os.path.isfile(candidate):
            pytesseract.pytesseract.tesseract_cmd = candidate
            try:
                pytesseract.pytesseract.get_tesseract_version()
                return True
            except Exception:
                continue
    return False


def _projection_variance_score(pil_img) -> float:
    """TẦNG 1: chấm điểm "độ rõ rệt các dòng ngang" của ảnh, KHÔNG dùng OCR.
    Ảnh đang ở đúng hướng đọc theo chiều ngang (chữ chạy dọc theo dòng) => variance CAO.
    Ảnh bị xoay 90°/270° (chữ chạy dọc theo cột) => variance THẤP."""
    gray = np.array(pil_img.convert("L"))
    threshold = gray.mean() - gray.std() * 0.5
    binary_text_mask = (gray < threshold).astype(np.uint8)
    row_sums = binary_text_mask.sum(axis=1)
    return float(row_sums.var())


def _ocr_confidence_score(img, lang: str = OCR_LANG) -> float:
    """TẦNG 2: chấm điểm 1 ảnh theo tổng độ tin cậy (confidence) OCR của các từ nhận diện
    được (lang=vie+eng). CHỈ dùng để phân biệt 2 góc đối lập nhau 180° (đã được TẦNG 1 xác
    định đúng nhóm), nên độ nhiễu watermark/dấu mộc không đủ để đánh lừa kết quả."""
    try:
        data = pytesseract.image_to_data(img, lang=lang, output_type=pytesseract.Output.DICT)
    except Exception:
        return 0.0
    total = 0.0
    for conf, text in zip(data.get("conf", []), data.get("text", [])):
        text = (text or "").strip()
        if not text:
            continue
        try:
            c = float(conf)
        except (TypeError, ValueError):
            c = -1
        if c > 0:
            total += c
    return total


def _detect_rotation_angle(pil_img, lang: str = OCR_LANG):
    """Áp dụng GIẢI PHÁP 2 TẦNG để xác định góc xoay đúng cho 1 ảnh trang:
        TẦNG 1: so sánh projection-variance ở 0° và 90°. CHỈ coi trang là "bị xoay ngang
                (nhóm {90°,270°})" nếu tỉ lệ variance(90°)/variance(0°) VƯỢT ngưỡng an toàn
                ROTATION_PROJECTION_RATIO_THRESHOLD; ngược lại (kể cả khi score_90 nhích cao
                hơn score_0 một chút do nhiễu ở trang văn bản thường/trang bìa) vẫn giữ nguyên
                nhóm {0°, 180°} để tránh xoay nhầm (đã kiểm chứng thực nghiệm: trang lỗi thật
                luôn có ratio >> ngưỡng, trang OK dù nhiễu cũng không vượt ngưỡng).
        TẦNG 2: dùng OCR confidence để chọn góc đúng CỤ THỂ trong 2 ứng viên của nhóm đã chọn.
    Trả về (góc_được_chọn, dict thông tin debug để ghi log)."""
    score_0 = _projection_variance_score(pil_img)
    rotated_90 = pil_img.rotate(-90, expand=True)
    score_90 = _projection_variance_score(rotated_90)

    ratio_90_over_0 = (score_90 / score_0) if score_0 > 0 else float("inf")
    if ratio_90_over_0 > ROTATION_PROJECTION_RATIO_THRESHOLD:
        candidates = (90, 270)
    else:
        candidates = (0, 180)

    ocr_scores = {}
    for angle in candidates:
        rotated = pil_img.rotate(-angle, expand=True) if angle != 0 else pil_img
        ocr_scores[angle] = _ocr_confidence_score(rotated, lang=lang)

    chosen_angle = max(ocr_scores, key=lambda a: ocr_scores[a])
    debug_info = {
        "projection_0": round(score_0, 1),
        "projection_90": round(score_90, 1),
        "ratio_90_over_0": round(ratio_90_over_0, 2),
        "candidates": candidates,
        "ocr_scores": ocr_scores,
    }
    return chosen_angle, debug_info



def normalize_pdf_orientation(pdf_path: str, log_callback) -> str:
    """GIẢI PHÁP CHÍNH: dò và sửa các trang PDF bị xoay sai hướng (ví dụ trang Báo cáo
    KQKD hợp nhất theo quý, do có 4 cột số liệu nên hay bị scan/in ngang rồi ghép vào file
    PDF chung mà không xoay lại đúng chiều đọc).

    Cách hoạt động (xem chi tiết ở phần chú giải TẦNG 1/TẦNG 2 phía trên):
        1) Render từng trang ra ảnh (PyMuPDF, DPI thấp chỉ để phát hiện hướng).
        2) TẦNG 1 (projection-variance): xác định nhóm góc đúng {0,180} hay {90,270}.
        3) TẦNG 2 (OCR confidence, lang=vie+eng): chọn góc cụ thể đúng trong nhóm đó.
        4) Nếu góc được chọn khác 0 => set cờ /Rotate cho đúng trang đó (không đổi ảnh gốc,
           chỉ đổi metadata hướng hiển thị khi render/xem/gửi đi).
        5) Nếu có ít nhất 1 trang được sửa, lưu ra 1 file PDF TẠM và trả về đường dẫn file
           tạm đó (file gốc của người dùng KHÔNG bị thay đổi). Nếu không có trang nào cần
           sửa, hoặc thiếu thư viện/Tesseract, trả về NGUYÊN đường dẫn file gốc.
    """
    if not (_ROTATION_FIX_CORE_LIBS_AVAILABLE and _configure_tesseract()):
        log_callback(
            "  !! Bỏ qua bước tự động phát hiện/sửa trang bị xoay ngang: thiếu thư viện "
            "PyMuPDF/numpy/pytesseract hoặc chưa cài Tesseract-OCR trên máy này. "
            "(pip install pymupdf numpy pytesseract, và cài đặt Tesseract-OCR để kích hoạt)."
        )
        return pdf_path

    log_callback("Đang kiểm tra hướng hiển thị của từng trang PDF (phát hiện trang bị xoay ngang)...")

    doc = fitz.open(pdf_path)
    changed_pages = []
    try:
        for page_index in range(len(doc)):
            page = doc[page_index]
            page.set_rotation(0)  # reset để lấy đúng nội dung gốc trước khi thử các góc
            pix = page.get_pixmap(dpi=OCR_DETECT_DPI)
            pil_img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            angle, debug_info = _detect_rotation_angle(pil_img)
            if angle != 0:
                page.set_rotation(angle)
                changed_pages.append((page_index + 1, angle))
                log_callback(
                    f"  -> Phát hiện trang {page_index + 1} bị xoay sai hướng, tự động xoay lại "
                    f"{angle}° (chi tiết: {debug_info})."
                )

        if not changed_pages:
            log_callback("  -> Không phát hiện trang nào bị xoay sai hướng.")
            return pdf_path

        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".pdf")
        os.close(tmp_fd)
        doc.save(tmp_path)
        log_callback(
            f"  -> Đã tự động sửa hướng cho {len(changed_pages)} trang "
            f"({', '.join(f'trang {p}: +{a}°' for p, a in changed_pages)}). "
            f"Đã lưu file PDF tạm để gửi trích xuất."
        )
        return tmp_path
    finally:
        doc.close()



# ============================== CẤU HÌNH ==============================


OPENROUTER_API_KEY = ""
OPENROUTER_MODEL = "openai/gpt-5.4-mini"
OPENROUTER_URL = "https://openrouter.ai/api/v1/chat/completions"

# Timeout dài vì PDF BCTC có thể nhiều trang, model cần thời gian xử lý
REQUEST_TIMEOUT_SECONDS = 600

MAX_QUEUE_SIZE = 10

APP_TITLE = "Trích xuất Báo cáo tài chính (OpenRouter · GPT-5.4-mini)"

# Các định dạng file phổ biến mà data analyst dùng để lưu dữ liệu báo cáo tài chính
FORMAT_OPTIONS = {
    "Excel (.xlsx)": "xlsx",
    "SQLite Database (.db)": "db",
    "CSV (3 file .csv)": "csv",
    "JSON (.json)": "json",
}
DEFAULT_FORMAT_LABEL = "Excel (.xlsx)"

REPORT_KEYS = [
    ("bang_can_doi_ke_toan", "Bảng cân đối kế toán", "can_doi_ke_toan", "Can doi ke toan"),
    ("bao_cao_ket_qua_hoat_dong_kinh_doanh", "Báo cáo kết quả hoạt động kinh doanh", "ket_qua_kinh_doanh", "Ket qua kinh doanh"),
    ("bao_cao_luu_chuyen_tien_te", "Báo cáo lưu chuyển tiền tệ", "luu_chuyen_tien_te", "Luu chuyen tien te"),
]
# tuple: (key trong JSON model trả về, tên hiển thị, tên gợi ý cho file/table csv/sqlite, tên sheet excel)

# Ghi chú (giải pháp chống thiếu dữ liệu LCTT/KQKD/BCĐKT):
# - Trước đây, chương trình gọi 1 request DUY NHẤT yêu cầu model trích xuất CẢ 3 báo cáo cùng lúc.
#   Với các doanh nghiệp có báo cáo dài (đặc biệt LCTT theo phương pháp trực tiếp trải trên nhiều
#   trang), tổng số token output cần sinh ra có thể vượt quá max_tokens => model buộc phải "cắt gọn"
#   nội dung (thường là báo cáo được sinh SAU CÙNG trong JSON, tức LCTT) để vẫn đóng được JSON hợp lệ,
#   dẫn đến thiếu dữ liệu mà không có cảnh báo nào.
# - Giải pháp áp dụng:
#     1) Tăng max_tokens cho mỗi lần gọi (xem MAX_TOKENS_PER_REPORT).
#     2) TÁCH thành 3 request riêng biệt, mỗi request chỉ yêu cầu trích xuất DUY NHẤT 1 báo cáo
#        (BCĐKT / KQKD / LCTT), sau đó ghép kết quả lại. Nhờ vậy mỗi báo cáo có toàn bộ ngân sách
#        token dành riêng cho nó, không còn báo cáo nào bị "hy sinh" vì hết token.
#     3) Kiểm tra finish_reason của response: nếu = "length" (bị cắt vì chạm max_tokens) thì báo lỗi
#        rõ ràng ngay lập tức, KHÔNG lặng lẽ xuất dữ liệu thiếu.

MAX_TOKENS_PER_REPORT = 32000

SINGLE_REPORT_SYSTEM_PROMPT_TEMPLATE = """Bạn là chuyên gia phân tích báo cáo tài chính doanh nghiệp Việt Nam.
Bạn sẽ đọc TRỰC TIẾP nội dung file PDF báo cáo tài chính được cung cấp (không có bước trung gian nào khác).

Nhiệm vụ DUY NHẤT của bạn: trích xuất chính xác TOÀN BỘ số liệu của MỘT báo cáo sau đây, giữ đúng
THỨ TỰ và THỂ THỨC trình bày như trong file PDF gốc:

    >>> {display_name} <<<

CẢNH BÁO QUAN TRỌNG: báo cáo này có thể trải dài trên NHIỀU TRANG của file PDF (kể cả các trang có
tiêu đề "... (tiếp theo)"). Bạn PHẢI đọc và trích xuất ĐẦY ĐỦ TẤT CẢ các trang có liên quan đến báo
cáo này, TUYỆT ĐỐI KHÔNG được bỏ sót phần nào (ví dụ nếu là Báo cáo lưu chuyển tiền tệ, phải có ĐỦ
CẢ 3 PHẦN: Lưu chuyển tiền từ hoạt động kinh doanh, hoạt động đầu tư, và hoạt động tài chính, cùng
số dư tiền đầu kỳ/cuối kỳ).

Yêu cầu bắt buộc:
- Giữ đúng thứ tự các chỉ tiêu (dòng) đúng như trong PDF, không tự ý sắp xếp lại.
- CHỈ trích xuất các dòng THỰC SỰ LÀ CHỈ TIÊU CÓ SỐ LIỆU (kể cả dòng tổng, dòng con, dòng
  "Trong đó" - miễn là dòng đó có mã số và/hoặc có ít nhất 1 giá trị số liệu đi kèm).
- TUYỆT ĐỐI KHÔNG tạo dòng riêng cho các nội dung KHÔNG PHẢI chỉ tiêu, bao gồm:
    a) Tiêu đề phân nhóm lớn không có mã số và không có số liệu đi kèm (ví dụ "TÀI SẢN",
       "NGUỒN VỐN") - đây chỉ là đề mục hiển thị, không mang thông tin số liệu, phải BỎ QUA.
    b) Dòng chú thích/diễn giải công thức tính toán, thường in nhỏ/nghiêng ngay dưới tên một
       chỉ tiêu, dạng "(100 = 110 + 120 + 130 + 140 + 150)" - đây KHÔNG phải một chỉ tiêu độc
       lập, phải BỎ QUA HOÀN TOÀN, KHÔNG tạo dòng riêng cho nó.
  Ví dụ minh hoạ: trong PDF gốc hiển thị liên tiếp:
      "Tài sản ngắn hạn
       (100 = 110 + 120 + 130 + 140 + 150)          100        63.943.158.727.826   56.747.258.197.010"
  thì đây CHỈ LÀ 1 (một) chỉ tiêu duy nhất, PHẢI trích xuất thành DUY NHẤT 1 dòng:
      {{"ma_so": "100", "chi_tieu": "Tài sản ngắn hạn", "thuyet_minh": "", "gia_tri": [63943158727826, 56747258197010]}}
  KHÔNG được tách thành 2 dòng (1 dòng "Tài sản ngắn hạn" không có mã số, và 1 dòng công thức
  "(100 = ...)" có mã số nhưng không có số liệu). Mã số 100 PHẢI được gán vào đúng dòng chỉ tiêu
  "Tài sản ngắn hạn" (dòng có số liệu thật), phần công thức diễn giải phải bị loại bỏ hoàn toàn.
- Lấy đúng "ma_so" (mã số) của từng chỉ tiêu nếu có trong báo cáo gốc.
- Lấy đúng "thuyet_minh" (số thuyết minh) nếu có.
- Lấy đúng toàn bộ các cột số liệu theo kỳ báo cáo có trong file. Đặt tên cột đúng theo tiêu đề cột
  trong file gốc.
- QUY TẮC BẮT BUỘC VỀ ĐỊNH DẠNG SỐ trong trường "gia_tri": mỗi giá trị PHẢI là kiểu SỐ (JSON number),
  KHÔNG PHẢI chuỗi (string), và PHẢI dùng dấu trừ chuẩn "-" để biểu diễn số âm, TUYỆT ĐỐI KHÔNG giữ
  lại dấu ngoặc "()" hoặc dấu phân cách hàng nghìn (dấu chấm/dấu phẩy) trong giá trị trả về.
  Trong bản gốc PDF, số âm được trình bày bằng dấu ngoặc - đây CHỈ là quy ước hiển thị, khi xuất
  JSON bạn PHẢI chuyển đổi thành số âm chuẩn. Ví dụ:
    - PDF hiển thị "(2.552.250.936)"  =>  JSON phải trả  -2552250936   (number, không phải chuỗi)
    - PDF hiển thị "1.454.908.359.540" =>  JSON phải trả  1454908359540 (number, không phải chuỗi)
  Áp dụng NHẤT QUÁN quy tắc này cho TẤT CẢ các giá trị số trong toàn bộ báo cáo.
- Nếu một chỉ tiêu không có số liệu (bỏ trống), để giá trị null.
- Không tự tính toán, không tự suy diễn số liệu không có trong file.
- Đơn vị tiền tệ: ghi lại đúng như trong file gốc (ví dụ VND, hoặc "triệu đồng"...).
- CHỈ trích xuất báo cáo "{display_name}" nêu trên. KHÔNG trích xuất các báo cáo tài chính khác
  trong file (nếu có).

Trả lời DUY NHẤT bằng một JSON object hợp lệ (không markdown, không giải thích thêm), theo đúng
cấu trúc sau:

{{
  "don_vi_tien_te": "<đơn vị tiền tệ ghi trong báo cáo>",
  "ky_bao_cao": "<kỳ báo cáo, ví dụ Quý I năm 2026>",
  "ten_doanh_nghiep": "<tên doanh nghiệp nếu có>",
  "{json_key}": {{
    "cot_du_lieu": ["<tên cột 1>", "<tên cột 2>", ...],
    "dong": [
      {{"ma_so": "<mã số hoặc rỗng>", "chi_tieu": "<tên chỉ tiêu>", "thuyet_minh": "<số thuyết minh hoặc rỗng>", "gia_tri": [<giá trị SỐ theo đúng thứ tự cot_du_lieu, số âm dùng dấu "-">]}},
      ...
    ]
  }}
}}

Chỉ trả về JSON, tuyệt đối không thêm bất kỳ văn bản nào khác trước hoặc sau JSON.
"""

SINGLE_REPORT_USER_PROMPT_TEMPLATE = (
    "Đây là file PDF báo cáo tài chính hợp nhất. Hãy đọc trực tiếp toàn bộ file PDF này và trích "
    "xuất ĐẦY ĐỦ, KHÔNG BỎ SÓT báo cáo \"{display_name}\" theo đúng định dạng JSON đã được mô tả "
    "trong system prompt. Nếu báo cáo này trải dài trên nhiều trang (kể cả trang '... (tiếp theo)'), "
    "hãy đọc và trích xuất HẾT tất cả các trang liên quan đến báo cáo này. Giữ nguyên thứ tự và mã số "
    "dòng như file gốc. CHỈ trích xuất các dòng chỉ tiêu có mã số và/hoặc số liệu thật; BỎ QUA hoàn "
    "toàn các dòng tiêu đề nhóm (như 'TÀI SẢN', 'NGUỒN VỐN') và các dòng công thức diễn giải kiểu "
    "'(100 = 110 + 120 + ...)'. Mọi giá trị số PHẢI ở dạng number, số âm dùng dấu '-', không dùng "
    "dấu ngoặc hay dấu phân cách hàng nghìn."
)




# ============================== LOGIC GỌI API ==============================

class ExtractionError(Exception):
    pass


def encode_pdf_to_base64(pdf_path: str) -> str:
    with open(pdf_path, "rb") as f:
        data = f.read()
    return base64.b64encode(data).decode("utf-8")


def _call_openrouter_single_report(
    b64_pdf: str, filename: str, json_key: str, display_name: str, log_callback
) -> dict:
    """Gửi 1 request tới OpenRouter yêu cầu trích xuất DUY NHẤT 1 báo cáo (json_key).
    Trả về dict JSON model trả về (chỉ chứa report đó + các meta field).
    Raise ExtractionError nếu request lỗi, hoặc bị cắt do hết max_tokens (finish_reason == 'length')."""

    system_prompt = SINGLE_REPORT_SYSTEM_PROMPT_TEMPLATE.format(
        display_name=display_name, json_key=json_key
    )
    user_prompt = SINGLE_REPORT_USER_PROMPT_TEMPLATE.format(display_name=display_name)

    payload = {
        "model": OPENROUTER_MODEL,
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": user_prompt},
                    {
                        "type": "file",
                        "file": {
                            "filename": filename,
                            "file_data": f"data:application/pdf;base64,{b64_pdf}",
                        },
                    },
                ],
            },
        ],
        # Cấu hình plugin xử lý PDF của OpenRouter: engine "native" để model đọc file PDF
        # trực tiếp, không qua OCR trung gian.
        "plugins": [
            {
                "id": "file-parser",
                "pdf": {"engine": "native"},
            }
        ],
        "temperature": 0,
        "max_tokens": MAX_TOKENS_PER_REPORT,
    }

    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json",
        "HTTP-Referer": "http://localhost",
        "X-Title": "BCTC Extractor",
    }

    log_callback(f"  -> Đang gửi request trích xuất riêng: {display_name} (max_tokens={MAX_TOKENS_PER_REPORT}) ...")

    try:
        resp = requests.post(
            OPENROUTER_URL,
            headers=headers,
            data=json.dumps(payload),
            timeout=REQUEST_TIMEOUT_SECONDS,
        )
    except requests.exceptions.RequestException as e:
        raise ExtractionError(f"LỖI KẾT NỐI ĐẾN OPENROUTER (báo cáo: {display_name}): {e}")

    if resp.status_code != 200:
        raise ExtractionError(
            f"LỖI TỪ OPENROUTER (báo cáo: {display_name}, HTTP {resp.status_code}): {resp.text[:3000]}"
        )

    try:
        resp_json = resp.json()
    except Exception as e:
        raise ExtractionError(
            f"LỖI PARSE JSON PHẢN HỒI TỪ OPENROUTER (báo cáo: {display_name}): {e}\n"
            f"Nội dung thô: {resp.text[:3000]}"
        )

    if "error" in resp_json:
        raise ExtractionError(f"LỖI API OPENROUTER (báo cáo: {display_name}): {resp_json['error']}")

    try:
        choice = resp_json["choices"][0]
        content = choice["message"]["content"]
    except (KeyError, IndexError) as e:
        raise ExtractionError(
            f"LỖI CẤU TRÚC PHẢN HỒI (báo cáo: {display_name}): không tìm thấy choices[0].message.content.\n"
            f"Chi tiết lỗi: {e}\nPhản hồi thô: {json.dumps(resp_json, ensure_ascii=False)[:3000]}"
        )

    # GIẢI PHÁP #3: kiểm tra finish_reason. Nếu model bị cắt vì hết max_tokens (finish_reason
    # == "length"), báo lỗi RÕ RÀNG ngay lập tức, không lặng lẽ export dữ liệu thiếu.
    finish_reason = choice.get("finish_reason") or choice.get("native_finish_reason")
    if finish_reason == "length":
        raise ExtractionError(
            f"LỖI: Phản hồi cho báo cáo \"{display_name}\" bị CẮT vì chạm giới hạn max_tokens "
            f"({MAX_TOKENS_PER_REPORT}) (finish_reason=length). Dữ liệu trả về CÓ THỂ BỊ THIẾU. "
            f"Hãy tăng MAX_TOKENS_PER_REPORT trong code hoặc thử lại."
        )

    log_callback(f"  -> Đã nhận phản hồi cho báo cáo: {display_name}. Đang phân tích JSON...")

    data = parse_json_from_model_output(content)
    return data


def call_openrouter_with_pdf(pdf_path: str, log_callback) -> dict:
    """GIẢI PHÁP #2: Gửi file PDF trực tiếp (base64) đến OpenRouter bằng 3 REQUEST RIÊNG BIỆT,
    mỗi request chỉ yêu cầu trích xuất DUY NHẤT 1 trong 3 báo cáo (BCĐKT / KQKD / LCTT), sau đó
    ghép kết quả lại thành 1 dict giống cấu trúc cũ để phần export không cần thay đổi.
    Nhờ tách request, mỗi báo cáo có toàn bộ ngân sách max_tokens riêng (GIẢI PHÁP #1: tăng
    max_tokens), tránh trường hợp báo cáo dài (đặc biệt LCTT) bị "hy sinh" vì hết token.

    GIẢI PHÁP #4 (chống đọc sai do trang bị xoay ngang): trước khi gửi PDF cho model, chạy
    normalize_pdf_orientation() để tự động phát hiện & sửa các trang có nội dung bị xoay sai
    hướng (ví dụ trang Báo cáo KQKD hợp nhất theo quý). Nếu có trang được sửa, dùng file PDF
    TẠM đã chuẩn hoá hướng để gửi đi (file PDF gốc của người dùng không bị thay đổi); file tạm
    này sẽ được dọn (xoá) ngay sau khi xử lý xong ở _run_queue_thread."""
    normalized_pdf_path = normalize_pdf_orientation(pdf_path, log_callback)
    # File PDF tạm (nếu có trang bị sửa hướng) chỉ được tạo khi normalized_pdf_path khác
    # pdf_path gốc. Phải dọn (xoá) file tạm này sau khi dùng xong, dù xử lý thành công hay lỗi.
    is_temp_pdf = normalized_pdf_path != pdf_path

    try:
        log_callback(f"Đang đọc file PDF: {normalized_pdf_path}")
        b64_pdf = encode_pdf_to_base64(normalized_pdf_path)
        filename = os.path.basename(pdf_path)

        log_callback(
            f"Sẽ gửi {len(REPORT_KEYS)} request riêng biệt tới OpenRouter (model: {OPENROUTER_MODEL}), "
            f"mỗi request trích xuất 1 báo cáo, để tránh bị cắt dữ liệu do giới hạn max_tokens."
        )
        log_callback("(Quá trình này có thể mất vài chục giây đến vài phút cho mỗi báo cáo)")

        merged_data = {}
        for json_key, display_name, _slug, _sheet_name in REPORT_KEYS:
            report_data = _call_openrouter_single_report(
                b64_pdf, filename, json_key, display_name, log_callback
            )

            # Gộp các field meta (đơn vị tiền tệ, kỳ báo cáo, tên DN) nếu chưa có
            for meta_key in ("don_vi_tien_te", "ky_bao_cao", "ten_doanh_nghiep"):
                if meta_key not in merged_data and report_data.get(meta_key):
                    merged_data[meta_key] = report_data.get(meta_key)

            report_content = report_data.get(json_key)
            if not report_content or not report_content.get("dong"):
                log_callback(
                    f"  !! CẢNH BÁO: báo cáo \"{display_name}\" trả về RỖNG hoặc không có dòng dữ liệu nào."
                )
            else:
                # GIẢI PHÁP #1 (bổ sung): lọc bỏ các dòng "rác" (tiêu đề nhóm / dòng công thức
                # diễn giải) mà model có thể vẫn lỡ trích xuất dư, dù đã được yêu cầu bỏ qua trong
                # prompt. Đây là lớp phòng thủ thứ 2, không phụ thuộc hoàn toàn vào việc model có
                # tuân thủ đúng prompt hay không.
                before_count = len(report_content.get("dong") or [])
                report_content["dong"] = _filter_junk_rows(report_content.get("dong") or [])
                after_count = len(report_content["dong"])
                if after_count < before_count:
                    log_callback(
                        f"  -> Đã lọc bỏ {before_count - after_count} dòng không hợp lệ "
                        f"(tiêu đề nhóm/dòng công thức) trong báo cáo \"{display_name}\"."
                    )
            merged_data[json_key] = report_content

        return merged_data
    finally:
        if is_temp_pdf and os.path.isfile(normalized_pdf_path):
            try:
                os.remove(normalized_pdf_path)
            except OSError:
                pass



# ============================== LỌC DÒNG "RÁC" (KHÔNG PHẢI CHỈ TIÊU) ==============================

# Regex nhận diện dòng công thức diễn giải dạng "(100 = 110 + 120 + 130 + 140 + 150)"
_FORMULA_ROW_RE = re.compile(r"^\(?\s*\d+\s*=\s*\d+(\s*[+\-]\s*\d+)*\s*\)?$")


def _is_junk_row(dong: dict) -> bool:
    """Nhận diện dòng KHÔNG phải chỉ tiêu thật (tiêu đề nhóm hoặc dòng công thức diễn giải),
    để loại bỏ khỏi kết quả cuối cùng, phòng trường hợp model vẫn lỡ trích xuất dư."""
    chi_tieu = str(dong.get("chi_tieu") or "").strip()
    ma_so = str(dong.get("ma_so") or "").strip()
    gia_tri = dong.get("gia_tri") or []
    has_any_value = any(v is not None and str(v).strip() != "" for v in gia_tri)

    # Dòng công thức diễn giải kiểu "(100 = 110 + 120 + ...)"
    chi_tieu_clean = chi_tieu.replace(" ", "")
    if _FORMULA_ROW_RE.match(chi_tieu_clean):
        return True

    # Dòng không có mã số VÀ không có bất kỳ giá trị số liệu nào => rất có thể là tiêu đề
    # nhóm (ví dụ "TÀI SẢN", "NGUỒN VỐN") chứ không phải 1 chỉ tiêu thật.
    if not ma_so and not has_any_value:
        return True

    return False


def _filter_junk_rows(dong_list: list) -> list:
    return [d for d in dong_list if not _is_junk_row(d)]




def parse_json_from_model_output(content: str) -> dict:
    """Model được yêu cầu trả JSON thuần, nhưng vẫn cố gắng bóc tách nếu có markdown code block."""
    if content is None:
        raise ExtractionError("LỖI: model trả về nội dung rỗng (content=None).")

    text = content.strip()

    fence_match = re.search(r"```(?:json)?\s*(\{.*\})\s*```", text, re.DOTALL)
    if fence_match:
        text = fence_match.group(1)
    else:
        first_brace = text.find("{")
        last_brace = text.rfind("}")
        if first_brace != -1 and last_brace != -1 and last_brace > first_brace:
            text = text[first_brace : last_brace + 1]

    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        raise ExtractionError(
            f"LỖI: Không parse được JSON từ phản hồi model.\n"
            f"Chi tiết lỗi JSONDecodeError: {e}\n"
            f"--- Nội dung model trả về (rút gọn 3000 ký tự đầu) ---\n{content[:3000]}"
        )


# ============================== TIỀN XỬ LÝ GIÁ TRỊ ==============================

def _sanitize_value(v):
    """Chuẩn hoá 1 giá trị số về kiểu int/float chuẩn, xử lý ĐÚNG các kiểu định dạng số
    thường gặp trong BCTC Việt Nam:
      - Dấu ngoặc "(...)" biểu thị số âm (ví dụ "(2.552.250.936)" => -2552250936).
      - Dấu chấm "." dùng làm dấu PHÂN CÁCH HÀNG NGHÌN (không phải dấu thập phân), ví dụ
        "2.552.250.936" => 2552250936. Chỉ coi dấu chấm là dấu thập phân khi nó là dấu chấm
        DUY NHẤT trong chuỗi và theo sau bởi 1-2 chữ số ở cuối (ví dụ "1234.5" => 1234.5).
      - Dấu phẩy "," dùng làm dấu phân cách hàng nghìn kiểu quốc tế (ví dụ "2,552,250,936").
    Đây là lớp phòng thủ CUỐI CÙNG, độc lập với model: dù model trả về number chuẩn hay
    chuỗi còn giữ nguyên định dạng gốc (dấu ngoặc/dấu chấm/dấu phẩy), hàm này đều phải cho
    ra kết quả số âm/dương thống nhất."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s in ("-", "—", "–"):
            return None

        neg = False
        s_clean = s.replace(" ", "")
        if s_clean.startswith("(") and s_clean.endswith(")"):
            neg = True
            s_clean = s_clean[1:-1]
        if s_clean.startswith("-"):
            neg = True
            s_clean = s_clean[1:]

        # Xác định dấu chấm cuối có phải dấu thập phân hay không (ví dụ "1234.5", "0.75").
        # Nếu chuỗi chỉ có 1 dấu chấm và phần sau dấu chấm cuối có 1-2 chữ số => dấu thập phân.
        # Ngược lại (nhiều dấu chấm, hoặc 3 chữ số sau dấu chấm) => dấu chấm là phân cách hàng nghìn.
        dot_count = s_clean.count(".")
        is_decimal_dot = False
        if dot_count == 1:
            after_dot = s_clean.rsplit(".", 1)[1]
            if 1 <= len(after_dot) <= 2 and after_dot.isdigit():
                is_decimal_dot = True

        if is_decimal_dot:
            s_num = s_clean.replace(",", "")
        else:
            # Toàn bộ dấu chấm và dấu phẩy đều là phân cách hàng nghìn => loại bỏ hết.
            s_num = s_clean.replace(".", "").replace(",", "")

        if s_num == "":
            return None

        try:
            num = float(s_num)
            if neg:
                num = -num
            if num == int(num):
                return int(num)
            return num
        except ValueError:
            return v
    return v



def _get_report(data: dict, key: str) -> dict:
    return data.get(key) or {"cot_du_lieu": [], "dong": []}


# ============================== XUẤT EXCEL (.xlsx) ==============================

HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)
BOLD_FONT = Font(bold=True)


def _write_excel_sheet(wb: Workbook, sheet_name: str, report: dict):
    """Chỉ ghi phần CỐT LÕI: dòng header (Mã số | Chỉ tiêu | Thuyết minh | ...cột số liệu)
    bắt đầu tại A1, và các dòng dữ liệu ngay sau đó. KHÔNG có tiêu đề báo cáo/tên DN/đơn vị."""
    ws = wb.create_sheet(title=sheet_name)

    cot_du_lieu = report.get("cot_du_lieu") or []
    dong_list = report.get("dong") or []

    header_row = 1
    headers = ["Mã số", "Chỉ tiêu", "Thuyết minh"] + list(cot_du_lieu)
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = header_row + 1
    for dong in dong_list:
        ma_so = dong.get("ma_so") or ""
        chi_tieu = dong.get("chi_tieu") or ""
        thuyet_minh = dong.get("thuyet_minh") or ""
        gia_tri = dong.get("gia_tri") or []

        ws.cell(row=r, column=1, value=ma_so).alignment = Alignment(horizontal="center")
        chi_tieu_cell = ws.cell(row=r, column=2, value=str(chi_tieu))
        ws.cell(row=r, column=3, value=thuyet_minh).alignment = Alignment(horizontal="center")

        for i in range(len(cot_du_lieu)):

            val = gia_tri[i] if i < len(gia_tri) else None
            val = _sanitize_value(val)
            vc = ws.cell(row=r, column=4 + i, value=val)
            if isinstance(val, (int, float)):
                vc.number_format = "#,##0"
                vc.alignment = Alignment(horizontal="right")
        r += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 10
    for i in range(len(cot_du_lieu)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 20

    ws.freeze_panes = f"D{header_row + 1}"


def export_xlsx(data: dict, output_path: str, log_callback) -> list:
    log_callback("Đang tạo file Excel (.xlsx)...")
    wb = Workbook()
    wb.remove(wb.active)

    for json_key, display_name, _slug, sheet_name in REPORT_KEYS:
        report = _get_report(data, json_key)
        _write_excel_sheet(wb, sheet_name, report)

    wb.save(output_path)
    log_callback(f"Đã lưu file: {output_path}")
    return [output_path]


# ============================== XUẤT SQLITE (.db) ==============================

def _sanitize_column_name(name: str, idx: int) -> str:
    name = (name or f"col_{idx}").strip()
    if not name:
        name = f"col_{idx}"
    return name


def export_db(data: dict, output_path: str, log_callback) -> list:
    log_callback("Đang tạo file SQLite Database (.db)...")
    if os.path.exists(output_path):
        os.remove(output_path)

    conn = sqlite3.connect(output_path)
    cur = conn.cursor()

    for json_key, display_name, slug, _sheet_name in REPORT_KEYS:
        report = _get_report(data, json_key)
        cot_du_lieu = report.get("cot_du_lieu") or []
        dong_list = report.get("dong") or []

        value_cols = [_sanitize_column_name(c, i) for i, c in enumerate(cot_du_lieu, start=1)]

        col_defs = ['"ma_so" TEXT', '"chi_tieu" TEXT', '"thuyet_minh" TEXT']
        for vc in value_cols:
            col_defs.append(f'"{vc}" REAL')

        table_name = slug
        cur.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        cur.execute(f'CREATE TABLE "{table_name}" ({", ".join(col_defs)})')

        placeholders = ", ".join(["?"] * (3 + len(value_cols)))
        insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'

        rows_to_insert = []
        for dong in dong_list:
            ma_so = dong.get("ma_so") or ""
            chi_tieu = dong.get("chi_tieu") or ""
            thuyet_minh = dong.get("thuyet_minh") or ""
            gia_tri = dong.get("gia_tri") or []

            row_values = [ma_so, chi_tieu, thuyet_minh]
            for i in range(len(value_cols)):
                val = gia_tri[i] if i < len(gia_tri) else None
                row_values.append(_sanitize_value(val))
            rows_to_insert.append(row_values)

        cur.executemany(insert_sql, rows_to_insert)
        log_callback(f"  - Đã tạo bảng '{table_name}' ({len(rows_to_insert)} dòng).")

    conn.commit()
    conn.close()
    log_callback(f"Đã lưu file: {output_path}")
    return [output_path]


# ============================== XUẤT JSON (.json) ==============================

def export_json(data: dict, output_path: str, log_callback) -> list:
    log_callback("Đang tạo file JSON (.json)...")
    result = {}
    for json_key, display_name, slug, _sheet_name in REPORT_KEYS:
        report = _get_report(data, json_key)
        cot_du_lieu = report.get("cot_du_lieu") or []
        dong_list = report.get("dong") or []

        clean_rows = []
        for dong in dong_list:
            gia_tri = dong.get("gia_tri") or []
            clean_rows.append({
                "ma_so": dong.get("ma_so") or "",
                "chi_tieu": dong.get("chi_tieu") or "",
                "thuyet_minh": dong.get("thuyet_minh") or "",
                "gia_tri": [_sanitize_value(v) for v in gia_tri],
            })

        result[slug] = {
            "cot_du_lieu": cot_du_lieu,
            "dong": clean_rows,
        }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    log_callback(f"Đã lưu file: {output_path}")
    return [output_path]


# ============================== XUẤT CSV (3 file .csv) ==============================

def export_csv(data: dict, output_dir: str, base_name_no_ext: str, log_callback) -> list:
    """CSV không hỗ trợ nhiều bảng trong 1 file => xuất 3 file riêng biệt.
    Tên file = <ten_pdf_goc>__<ten_bang>.csv để vẫn giữ tên gốc làm tiền tố."""
    log_callback("Đang tạo 3 file CSV (mỗi báo cáo 1 file)...")
    output_paths = []

    for json_key, display_name, slug, _sheet_name in REPORT_KEYS:
        report = _get_report(data, json_key)
        cot_du_lieu = report.get("cot_du_lieu") or []
        dong_list = report.get("dong") or []

        out_path = os.path.join(output_dir, f"{base_name_no_ext}__{slug}.csv")
        with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            headers = ["Ma so", "Chi tieu", "Thuyet minh"] + list(cot_du_lieu)
            writer.writerow(headers)
            for dong in dong_list:
                ma_so = dong.get("ma_so") or ""
                chi_tieu = dong.get("chi_tieu") or ""
                thuyet_minh = dong.get("thuyet_minh") or ""
                gia_tri = dong.get("gia_tri") or []
                row = [ma_so, chi_tieu, thuyet_minh]
                for i in range(len(cot_du_lieu)):
                    val = gia_tri[i] if i < len(gia_tri) else None
                    row.append(_sanitize_value(val))
                writer.writerow(row)

        log_callback(f"  - Đã lưu: {out_path}")
        output_paths.append(out_path)

    return output_paths


# ============================== ĐIỀU HƯỚNG XUẤT FILE ==============================

def export_data(data: dict, output_dir: str, base_name_no_ext: str, fmt: str, log_callback) -> list:
    """fmt: 'xlsx' | 'db' | 'json' | 'csv'"""
    if fmt == "xlsx":
        return export_xlsx(data, os.path.join(output_dir, base_name_no_ext + ".xlsx"), log_callback)
    elif fmt == "db":
        return export_db(data, os.path.join(output_dir, base_name_no_ext + ".db"), log_callback)
    elif fmt == "json":
        return export_json(data, os.path.join(output_dir, base_name_no_ext + ".json"), log_callback)
    elif fmt == "csv":
        return export_csv(data, output_dir, base_name_no_ext, log_callback)
    else:
        raise ExtractionError(f"LỖI: định dạng xuất không được hỗ trợ: {fmt}")


# ============================== GIAO DIỆN TKINTER ==============================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("760x600")
        self.resizable(True, True)

        self.out_dir_var = tk.StringVar()
        self.format_var = tk.StringVar(value=DEFAULT_FORMAT_LABEL)
        self.pdf_queue = []  # list of absolute paths

        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 8, "pady": 6}

        # ---- Khu vực hàng đợi PDF ----
        queue_frame = ttk.LabelFrame(self, text=f"Hàng đợi file PDF (tối đa {MAX_QUEUE_SIZE} file)")
        queue_frame.pack(fill="both", expand=False, **pad)

        list_container = ttk.Frame(queue_frame)
        list_container.pack(fill="both", expand=True, padx=6, pady=6)

        self.queue_listbox = tk.Listbox(list_container, height=8, selectmode="extended")
        self.queue_listbox.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(list_container, orient="vertical", command=self.queue_listbox.yview)
        scrollbar.pack(side="left", fill="y")
        self.queue_listbox.config(yscrollcommand=scrollbar.set)

        btns_frame = ttk.Frame(list_container)
        btns_frame.pack(side="left", fill="y", padx=(8, 0))
        ttk.Button(btns_frame, text="Thêm PDF...", command=self.add_pdfs).pack(fill="x", pady=2)
        ttk.Button(btns_frame, text="Xóa mục chọn", command=self.remove_selected).pack(fill="x", pady=2)
        ttk.Button(btns_frame, text="Xóa tất cả", command=self.clear_queue).pack(fill="x", pady=2)

        # ---- Thư mục lưu + định dạng ----
        opts_frame = ttk.Frame(self)
        opts_frame.pack(fill="x", **pad)

        ttk.Label(opts_frame, text="Thư mục lưu kết quả:").grid(row=0, column=0, sticky="w")
        ttk.Entry(opts_frame, textvariable=self.out_dir_var, width=55).grid(row=1, column=0, sticky="we")
        ttk.Button(opts_frame, text="Chọn thư mục...", command=self.choose_out_dir).grid(row=1, column=1, padx=4)

        ttk.Label(opts_frame, text="Định dạng file lưu trữ:").grid(row=2, column=0, sticky="w", pady=(10, 0))
        self.format_combo = ttk.Combobox(
            opts_frame, textvariable=self.format_var, values=list(FORMAT_OPTIONS.keys()),
            state="readonly", width=30,
        )
        self.format_combo.grid(row=3, column=0, sticky="w")

        opts_frame.columnconfigure(0, weight=1)

        # ---- Nút xử lý ----
        self.extract_btn = ttk.Button(self, text="Bắt đầu xử lý hàng đợi", command=self.on_start_clicked)
        self.extract_btn.pack(pady=6)

        # ---- Progress ----
        self.progress = ttk.Progressbar(self, mode="determinate", maximum=100)
        self.progress.pack(fill="x", padx=8, pady=(0, 6))
        self.progress_label_var = tk.StringVar(value="")
        ttk.Label(self, textvariable=self.progress_label_var).pack(anchor="w", padx=8)

        # ---- Log ----
        ttk.Label(self, text="Log tiến trình / Lỗi:").pack(anchor="w", padx=8)
        self.log_box = scrolledtext.ScrolledText(self, height=14, wrap="word")
        self.log_box.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    # -------------------- Quản lý hàng đợi --------------------

    def add_pdfs(self):
        paths = filedialog.askopenfilenames(
            title="Chọn file PDF báo cáo tài chính",
            filetypes=[("PDF files", "*.pdf")],
        )
        if not paths:
            return

        added, skipped_dup, skipped_limit = 0, 0, 0
        for p in paths:
            if len(self.pdf_queue) >= MAX_QUEUE_SIZE:
                skipped_limit += 1
                continue
            if p in self.pdf_queue:
                skipped_dup += 1
                continue
            self.pdf_queue.append(p)
            self.queue_listbox.insert("end", os.path.basename(p))
            added += 1

        if skipped_limit:
            messagebox.showwarning(
                "Vượt quá giới hạn",
                f"Chỉ được thêm tối đa {MAX_QUEUE_SIZE} file PDF. Đã bỏ qua {skipped_limit} file vượt giới hạn.",
            )
        if skipped_dup:
            self.log(f"Đã bỏ qua {skipped_dup} file trùng lặp trong hàng đợi.")

    def remove_selected(self):
        selected_indices = list(self.queue_listbox.curselection())
        if not selected_indices:
            return
        for idx in reversed(selected_indices):
            self.queue_listbox.delete(idx)
            del self.pdf_queue[idx]

    def clear_queue(self):
        self.queue_listbox.delete(0, "end")
        self.pdf_queue = []

    # -------------------- Thư mục / định dạng --------------------

    def choose_out_dir(self):
        path = filedialog.askdirectory(title="Chọn thư mục lưu kết quả")
        if path:
            self.out_dir_var.set(path)

    # -------------------- Log / trạng thái --------------------

    def log(self, message: str):
        def _append():
            timestamp = datetime.now().strftime("%H:%M:%S")
            self.log_box.insert("end", f"[{timestamp}] {message}\n")
            self.log_box.see("end")
        self.after(0, _append)

    def set_progress(self, current: int, total: int, label: str = ""):
        def _set():
            pct = 0 if total == 0 else int(current / total * 100)
            self.progress["value"] = pct
            self.progress_label_var.set(label)
        self.after(0, _set)

    def set_busy(self, busy: bool):
        def _set():
            state = "disabled" if busy else "normal"
            self.extract_btn.config(state=state)
            self.format_combo.config(state="disabled" if busy else "readonly")
        self.after(0, _set)

    # -------------------- Xử lý chính --------------------

    def on_start_clicked(self):
        if not self.pdf_queue:
            messagebox.showerror("Thiếu thông tin", "Vui lòng thêm ít nhất 1 file PDF vào hàng đợi.")
            return

        out_dir = self.out_dir_var.get().strip()
        if not out_dir:
            messagebox.showerror("Thiếu thông tin", "Vui lòng chọn thư mục lưu kết quả.")
            return
        if not os.path.isdir(out_dir):
            messagebox.showerror("Lỗi", f"Thư mục không tồn tại:\n{out_dir}")
            return

        format_label = self.format_var.get()
        fmt = FORMAT_OPTIONS.get(format_label)
        if not fmt:
            messagebox.showerror("Lỗi", "Vui lòng chọn định dạng file hợp lệ.")
            return

        self.log_box.delete("1.0", "end")
        self.set_progress(0, len(self.pdf_queue), "")
        self.set_busy(True)

        queue_copy = list(self.pdf_queue)
        thread = threading.Thread(
            target=self._run_queue_thread, args=(queue_copy, out_dir, fmt), daemon=True
        )
        thread.start()

    def _run_queue_thread(self, queue_copy, out_dir, fmt):
        total = len(queue_copy)
        success_count = 0
        fail_count = 0

        for i, pdf_path in enumerate(queue_copy, start=1):
            base_name_no_ext = os.path.splitext(os.path.basename(pdf_path))[0]
            self.set_progress(i - 1, total, f"Đang xử lý file {i}/{total}: {os.path.basename(pdf_path)}")
            self.log(f"===== [{i}/{total}] Bắt đầu xử lý: {pdf_path} =====")
            try:
                data = call_openrouter_with_pdf(pdf_path, self.log)
                output_paths = export_data(data, out_dir, base_name_no_ext, fmt, self.log)
                for p in output_paths:
                    self.log(f"[{i}/{total}] ĐÃ LƯU: {p}")
                self.log(f"[{i}/{total}] HOÀN THÀNH file: {os.path.basename(pdf_path)}")
                success_count += 1
            except ExtractionError as e:
                fail_count += 1
                self.log(f"[{i}/{total}] === LỖI (file: {os.path.basename(pdf_path)}) ===")
                self.log(str(e))
            except Exception:
                fail_count += 1
                tb = traceback.format_exc()
                self.log(f"[{i}/{total}] === LỖI KHÔNG XÁC ĐỊNH (file: {os.path.basename(pdf_path)}) ===")
                self.log(tb)

            self.set_progress(i, total, f"Đã xử lý {i}/{total} file")

        self.log("===== KẾT THÚC HÀNG ĐỢI =====")
        self.log(f"Thành công: {success_count} | Lỗi: {fail_count} | Tổng: {total}")
        self.set_busy(False)

        def _notify():
            if fail_count == 0:
                messagebox.showinfo("Hoàn thành", f"Đã xử lý xong {success_count}/{total} file thành công.")
            else:
                messagebox.showwarning(
                    "Hoàn thành (có lỗi)",
                    f"Đã xử lý xong. Thành công: {success_count}/{total}. Lỗi: {fail_count}.\n"
                    f"Xem chi tiết lỗi trong khung Log.",
                )
        self.after(0, _notify)


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
